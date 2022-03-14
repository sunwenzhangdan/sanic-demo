import logging

import openpyxl

from db.models import Users, OutboundStatistics
from sanic import Sanic, response

from tortoise.contrib.sanic import register_tortoise

logging.basicConfig(level=logging.DEBUG)

app = Sanic(__name__)


@app.get("/user")
async def list_all(request):
    # book = await openpyxl.load_workbook('./oc.xlsx')
    # sheet = book.get_sheet_by_name('出库统计表')
    # print("*******",sheet)
    # print("***********************************")
    users = await Users.all()
    return response.json({"users": [str(user) for user in users]})


@app.post("/user")
async def add_user(request):
    # book = await openpyxl.load_workbook('./oc.xlsx')
    # sheet = book.active
    # a1 = sheet['出库统计表']
    # print("*******",a1)
    print("***********************************")
    user = await Users.create(**request.json)
    return response.json({"user": str(user)})


@app.get("/statics")
async def add_statics(request):
    wb = openpyxl.load_workbook('oc.xlsx', read_only=True, data_only=True)
    sh = wb["出库统计表"]
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        print(row)
        stat = {"sap_no": row[0], "out_type": row[1], "out_category": row[2], "deliver_year ": row[3],
                "mark_month": row[4], "out_month": row[5], "out_date": row[6], "province ": row[7], "city": row[8],
                "area": row[9], "area_level": row[10], "bus_owner": row[11],
                "customer_name ": row[12], "product_name": row[13], "package_form": row[14],
                "specifications": row[15], "out_num": row[16], "single_price": row[17], "sale_num": row[18],
                "out_no": row[19], "deliver_no": row[20], "comment": row[21], "batch_no": row[22],
                "dead_line": row[23], "return_label": row[24], "invoice_pass": row[25], "invoice_type ": row[26],
                "invoice_count  ": row[27], "promoter_company_name": row[28], "foregin_money": row[29],
                "crm_client_code": row[30], "pass_one": row[31], "pass_two": row[32], "pass_unit": row[33],
                "out_num_discount": row[34]}
        if row[0] == "SAP订单号":
            continue
        await OutboundStatistics.create(**stat)
    return response.json({"users": "hello"})


@app.get("/statics/all")
async def list_all(request):
    statics = await OutboundStatistics.all()
    print(statics)
    for i in statics:
        d = str(i)
        print(d)
    print(type(statics))
    return response.json(statics)


register_tortoise(
    app, db_url="mysql://root:sun890816@localhost/rust", modules={"models": ["db.models"]}, generate_schemas=True
)

if __name__ == '__main__':
    import uvicorn

    uvicorn.run('main:app', host='0.0.0.0', port=8000, debug=True)
